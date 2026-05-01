"""
Conciliação Contajur
====================
Fluxo:
  1. parse_extrato()       → lista no formato padrão
  2. parse_base_extrato()  → converte base contábil para o mesmo formato
  3. reconcile()           → cruza os dois
  4. gerar_excel()         → Excel com 5 abas (análise)
  5. gerar_csv_final()     → CSV extrato_pendente no formato extrator

Formato padrão (dict):
  { data, descricao, valor_str, valor_num, tipo }
"""

import re
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# Helpers de valor
# ──────────────────────────────────────────────────────────────────────────────

def normalize_value(s: str) -> float:
    """'1.066,25' → 1066.25  |  '25.173' → 25173.0  |  '101' → 101.0"""
    s = s.strip().replace('\xa0', '').replace(' ', '')
    if not s:
        return 0.0
    if ',' in s:
        s = s.replace('.', '').replace(',', '.')
    else:
        s = s.replace('.', '')
    try:
        return float(s)
    except ValueError:
        return 0.0


def format_value_br(f: float) -> str:
    """1066.25 → '1.066,25'  |  25173.0 → '25.173'"""
    f = abs(f)
    cents = round(f * 100)
    int_part = cents // 100
    dec_part = cents % 100
    if dec_part == 0:
        return f"{int_part:,}".replace(",", ".")
    return f"{int_part:,}".replace(",", ".") + f",{dec_part:02d}"


# ──────────────────────────────────────────────────────────────────────────────
# Parsing
# ──────────────────────────────────────────────────────────────────────────────

DATE_RE = re.compile(r'^\d{2}/\d{2}/\d{4}$')


def _make_row(data, descricao, valor_str, tipo) -> dict:
    return {
        'data':      data,
        'descricao': descricao,
        'valor_str': valor_str,
        'valor_num': normalize_value(valor_str),
        'tipo':      tipo,
    }


def parse_extrato(file_bytes: bytes) -> list[dict]:
    """CSV sem cabeçalho: Data;Descrição;Valor;Tipo"""
    rows = []
    text = file_bytes.decode('utf-8-sig', errors='replace')
    for line in text.splitlines():
        parts = line.strip().split(';')
        if len(parts) < 4:
            continue
        data, desc, valor_str, tipo = parts[0], parts[1], parts[2], parts[3]
        if not DATE_RE.match(data.strip()):
            continue
        tipo = tipo.strip().upper()
        if tipo not in ('C', 'D'):
            continue
        rows.append(_make_row(data.strip(), desc.strip(), valor_str.strip(), tipo))
    return rows


def _extract_desc(historico: str) -> str:
    """
    'Recebimento 610 - POUSADA E RESTAURANTE LTDA, ref 12/2025'
     → 'POUSADA E RESTAURANTE LTDA ref 12/2025'
    """
    m = re.match(r'(?:Recebimento|Pagamento)\s+\S+\s*-\s*(.+)', historico, re.IGNORECASE)
    if m:
        desc = m.group(1).strip()
        desc = re.sub(r',\s*ref\s+', ' ref ', desc, flags=re.IGNORECASE)
        return desc
    return historico


def parse_base_extrato(file_bytes: bytes) -> list[dict]:
    """Base contábil (Razão/Livro Caixa) → formato extrator."""
    rows = []
    text = file_bytes.decode('utf-8-sig', errors='replace')
    current_date = ''

    for line in text.splitlines():
        parts = line.strip().split(';')
        if not parts or all(p.strip() == '' for p in parts):
            continue

        historico = parts[0].strip()

        if DATE_RE.match(historico) and all(p.strip() == '' for p in parts[1:]):
            current_date = historico
            continue

        if re.match(r'hist', historico, re.IGNORECASE):
            continue
        if re.match(r'\d+\s*-\s*BANCOS', historico, re.IGNORECASE):
            continue
        if 'saldo anterior' in historico.lower():
            continue

        debito  = parts[2].strip() if len(parts) > 2 else ''
        credito = parts[3].strip() if len(parts) > 3 else ''

        if debito:
            valor_str, tipo = debito, 'D'
        elif credito:
            valor_str, tipo = credito, 'C'
        else:
            continue

        if normalize_value(valor_str) == 0:
            continue

        rows.append(_make_row(current_date, _extract_desc(historico), valor_str, tipo))

    return rows


# ──────────────────────────────────────────────────────────────────────────────
# Similaridade de nome
# ──────────────────────────────────────────────────────────────────────────────

_STOPWORDS = {
    'DE', 'DA', 'DO', 'DAS', 'DOS', 'E', 'A', 'O', 'EM',
    'REF', 'PIX', 'PARA', 'COM', 'POR', 'NA', 'NO',
    'OUTRA', 'RECEBIMENTO', 'PAGAMENTO', 'SICOOB', 'IF',
}


def _keywords(text: str) -> set[str]:
    words = re.split(r'[\s.,;/\-_()0-9]+', text.upper())
    return {w for w in words if len(w) >= 4 and w not in _STOPWORDS}


def name_match(desc_base: str, desc_extrato: str, min_words: int = 2) -> bool:
    kw = _keywords(desc_base)
    if not kw:
        return False
    desc_up = desc_extrato.upper()
    hits = sum(1 for w in kw if w in desc_up)
    return hits >= min(min_words, len(kw))


# ──────────────────────────────────────────────────────────────────────────────
# Conciliação
# ──────────────────────────────────────────────────────────────────────────────

def reconcile(extrato: list[dict], base: list[dict]):
    """
    Retorna (conciliados, diferencas, so_extrato, so_base).

    Critérios (em ordem de prioridade):
      Conciliado : mesma data + mesmo tipo + mesmo valor
      Diferença  : mesma data + nome bate (valor diferente → mostra a diferença)
    """
    conciliados = []
    diferencas  = []
    ext_used    = set()
    base_used   = set()

    # Pass 1 – mesma data + mesmo valor (tipo é invertido: C no extrato = D na base)
    for i, b in enumerate(base):
        for j, e in enumerate(extrato):
            if j in ext_used:
                continue
            if (b['data'] == e['data']
                    and b['tipo'] != e['tipo']
                    and abs(b['valor_num'] - e['valor_num']) < 0.02):
                conciliados.append({'base': b, 'extrato': e})
                base_used.add(i)
                ext_used.add(j)
                break

    # Pass 2 – mesma data + nome bate (valor diferente)
    for i, b in enumerate(base):
        if i in base_used:
            continue
        for j, e in enumerate(extrato):
            if j in ext_used:
                continue
            if b['data'] == e['data'] and name_match(b['descricao'], e['descricao']):
                diff = abs(e['valor_num'] - b['valor_num'])
                diferencas.append({'base': b, 'extrato': e, 'diferenca': diff})
                base_used.add(i)
                ext_used.add(j)
                break

    so_extrato = [e for j, e in enumerate(extrato) if j not in ext_used]
    so_base    = [b for i, b in enumerate(base)    if i not in base_used]

    return conciliados, diferencas, so_extrato, so_base


# ──────────────────────────────────────────────────────────────────────────────
# Geração do Excel
# ──────────────────────────────────────────────────────────────────────────────

_GREEN  = PatternFill('solid', fgColor='C6EFCE')
_RED    = PatternFill('solid', fgColor='FFC7CE')
_YELLOW = PatternFill('solid', fgColor='FFEB9C')
_BLUE   = PatternFill('solid', fgColor='BDD7EE')
_GREY   = PatternFill('solid', fgColor='D9D9D9')
_BOLD   = Font(bold=True)
_CENTER = Alignment(horizontal='center')
_THIN   = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin'),
)

_EXT_COLS  = ['Data', 'Descrição', 'Valor', 'Tipo']
_BASE_COLS = ['Data (Base)', 'Descrição (Base)', 'Valor (Base)', 'Tipo (Base)']


def _header(ws, cols, fill):
    ws.append(cols)
    for cell in ws[ws.max_row]:
        cell.font = _BOLD; cell.fill = fill
        cell.alignment = _CENTER; cell.border = _THIN


def _style(ws, fill=None):
    for cell in ws[ws.max_row]:
        if fill: cell.fill = fill
        cell.border = _THIN


def _auto_width(ws):
    for col in ws.columns:
        w = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 70)


def _fmt(r: dict) -> list:
    return [r['data'], r['descricao'], r['valor_str'], r['tipo']]


def gerar_excel(conciliados, diferencas, so_extrato, so_base) -> io.BytesIO:
    wb = Workbook()

    # ── Aba 1: Conciliados ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Conciliados'
    _header(ws1, _EXT_COLS + _BASE_COLS, _GREEN)
    for c in conciliados:
        ws1.append(_fmt(c['extrato']) + _fmt(c['base']))
        _style(ws1, _GREEN)
    _auto_width(ws1)

    # ── Aba 2: Diferenças de Valor ────────────────────────────────────────────
    ws2 = wb.create_sheet('Diferenças de Valor')
    _header(ws2, _EXT_COLS + _BASE_COLS + ['Diferença'], _YELLOW)
    for d in diferencas:
        ws2.append(_fmt(d['extrato']) + _fmt(d['base']) + [format_value_br(d['diferenca'])])
        _style(ws2, _YELLOW)
    _auto_width(ws2)

    # ── Aba 3: Só no Extrato ──────────────────────────────────────────────────
    ws3 = wb.create_sheet('Só no Extrato')
    _header(ws3, _EXT_COLS, _RED)
    for e in so_extrato:
        ws3.append(_fmt(e))
        _style(ws3, _RED)
    _auto_width(ws3)

    # ── Aba 4: Só na Base ─────────────────────────────────────────────────────
    ws4 = wb.create_sheet('Só na Base')
    _header(ws4, _EXT_COLS, _BLUE)
    for b in so_base:
        ws4.append(_fmt(b))
        _style(ws4, _BLUE)
    _auto_width(ws4)

    # ── Aba 5: Resumo ─────────────────────────────────────────────────────────
    ws5 = wb.create_sheet('Resumo')
    _header(ws5, ['Categoria', 'Qtd', 'Total (Extrato)'], _GREY)

    def _tot(lst):
        return format_value_br(sum(r['valor_num'] for r in lst))

    ws5.append(['Conciliados',         len(conciliados), _tot([c['extrato'] for c in conciliados])])
    ws5.append(['Diferenças de Valor', len(diferencas),  _tot([d['extrato'] for d in diferencas])])
    ws5.append(['Só no Extrato',       len(so_extrato),  _tot(so_extrato)])
    ws5.append(['Só na Base',          len(so_base),     _tot(so_base)])
    _auto_width(ws5)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ──────────────────────────────────────────────────────────────────────────────
# CSV final no formato extrator
# ──────────────────────────────────────────────────────────────────────────────

def gerar_csv_final(diferencas: list[dict], so_extrato: list[dict]) -> io.BytesIO:
    """
    Extrato pendente (Data;Descrição;Valor;Tipo sem cabeçalho):
    - Diferenças → valor líquido |extrato − base|
    - Só no extrato → valor original
    """
    lines = []
    for d in diferencas:
        e    = d['extrato']
        diff = abs(d['diferenca'])
        if diff < 0.01:
            continue
        lines.append(f"{e['data']};{e['descricao']};{format_value_br(diff)};{e['tipo']}")
    for e in so_extrato:
        lines.append(f"{e['data']};{e['descricao']};{e['valor_str']};{e['tipo']}")

    buf = io.BytesIO('\n'.join(lines).encode('utf-8-sig'))
    buf.seek(0)
    return buf


# ──────────────────────────────────────────────────────────────────────────────
# Entry point
# ──────────────────────────────────────────────────────────────────────────────

def processar(extrato_bytes: bytes, base_bytes: bytes) -> tuple[io.BytesIO, io.BytesIO]:
    extrato = parse_extrato(extrato_bytes)
    base    = parse_base_extrato(base_bytes)

    if not extrato:
        raise ValueError("Nenhuma transação encontrada no extrato.")
    if not base:
        raise ValueError("Nenhuma transação encontrada na base contábil.")

    conciliados, diferencas, so_extrato, so_base = reconcile(extrato, base)
    excel_buf = gerar_excel(conciliados, diferencas, so_extrato, so_base)
    csv_buf   = gerar_csv_final(diferencas, so_extrato)

    return excel_buf, csv_buf
