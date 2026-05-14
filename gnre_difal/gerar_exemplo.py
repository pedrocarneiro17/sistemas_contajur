"""
Gera modelo_difal.xlsx — planilha em branco com todas as colunas GNRE DIFAL.
Execute:  python gerar_exemplo.py
"""

from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Definição das colunas
# ---------------------------------------------------------------------------

COLUNAS = [
    # (nome_coluna,           obrigatório, descrição/exemplo)
    ("uf_favorecida",         True,  "UF destinatária — ex: SP"),
    ("cnpj_emitente",         True,  "CNPJ da empresa (só números) — ex: 09595247000172"),
    ("razao_social",          True,  "Razão social da empresa"),
    ("uf_emitente",           True,  "UF da empresa — ex: SP"),
    ("municipio_emitente",    True,  "Código IBGE 5 dígitos — ex: 50308"),
    ("periodo_referencia",    True,  "Competência AAAAMM — ex: 202501"),
    ("valor_principal",       True,  "Valor ICMS DIFAL — ex: 1500.00"),
    ("data_vencimento",       True,  "Vencimento AAAA-MM-DD — ex: 2025-02-20"),
    ("tipo_gnre",             False, "0=Simples (padrão) | 1=Múlt. docs | 2=Múlt. receitas"),
    ("receita",               False, "Código 6 dígitos — vazio = 100080 (DIFAL padrão)"),
    ("detalhamento_receita",  False, "Código exigido por certas UFs — consulte o portal"),
    ("valor_fecp",            False, "Valor FCP/FECP — preencher para AL, BA, CE, MA, MG, PB, PE, PI, RJ, SE"),
    ("valor_juros",           False, "Juros — se houver"),
    ("valor_multa",           False, "Multa — se houver"),
    ("data_pagamento",        False, "Data pagamento AAAA-MM-DD — vazio = usa data_vencimento"),
    ("endereco",              False, "Endereço do emitente"),
    ("cep",                   False, "CEP só números — ex: 01310100"),
    ("telefone",              False, "Telefone só números — ex: 11999999999"),
    ("ie_emitente",           False, "IE do emitente na UF favorecida (se inscrito)"),
    ("cnpj_destinatario",     False, "CNPJ do destinatário (quando exigido pela UF)"),
    ("cpf_destinatario",      False, "CPF do destinatário (quando exigido pela UF)"),
    ("ie_destinatario",       False, "IE do destinatário"),
    ("razao_social_destinatario", False, "Razão social do destinatário"),
    ("municipio_destinatario",False, "Código IBGE 5 dígitos do destinatário"),
    ("documento_origem_tipo", False, "Tipo do doc: 55=NF-e | 57=CT-e | 65=NFC-e | 10=NF papel"),
    ("documento_origem",      False, "Chave/número do documento de origem"),
    ("identificador_guia",    False, "ID interno (até 10 dígitos) — para controle próprio"),
]

CABECALHOS  = [c[0] for c in COLUNAS]
OBRIGATORIO = [c[1] for c in COLUNAS]
DESCRICOES  = [c[2] for c in COLUNAS]

# ---------------------------------------------------------------------------
# Geração do arquivo
# ---------------------------------------------------------------------------

saida = Path("modelo_difal.xlsx")

df = pd.DataFrame(columns=CABECALHOS)
df.to_excel(saida, index=False, sheet_name="DIFAL")

wb = load_workbook(saida)
ws = wb["DIFAL"]

# --- Estilos ---
cor_obrig  = "1F4E79"   # azul escuro  → obrigatório
cor_opcio  = "2E75B6"   # azul médio   → opcional
cor_header = "FFFFFF"   # texto branco
cor_desc   = "D9E1F2"   # fundo azul claro para linha de descrição

borda_fina = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

# Linha 1 — cabeçalhos
for col_idx, (nome, obrig, _) in enumerate(COLUNAS, start=1):
    cell = ws.cell(row=1, column=col_idx, value=nome)
    cor_fundo = cor_obrig if obrig else cor_opcio
    cell.fill      = PatternFill("solid", fgColor=cor_fundo)
    cell.font      = Font(bold=True, color=cor_header, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = borda_fina

# Linha 2 — descrições (serve como guia)
for col_idx, (_, obrig, descricao) in enumerate(COLUNAS, start=1):
    cell = ws.cell(row=2, column=col_idx, value=descricao)
    cell.fill      = PatternFill("solid", fgColor=cor_desc)
    cell.font      = Font(italic=True, color="404040", size=9)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border    = borda_fina

# Ajusta largura das colunas
for col_idx, (nome, _, descricao) in enumerate(COLUNAS, start=1):
    tamanho = max(len(nome), len(descricao) // 2, 14)
    ws.column_dimensions[get_column_letter(col_idx)].width = min(tamanho, 40)

# Congela a linha de cabeçalho
ws.freeze_panes = "A3"

# Altura das linhas
ws.row_dimensions[1].height = 28
ws.row_dimensions[2].height = 36

# ---------------------------------------------------------------------------
# Aba de instruções
# ---------------------------------------------------------------------------

wi = wb.create_sheet("Instruções")
instrucoes = [
    ("INSTRUÇÕES DE PREENCHIMENTO", True),
    ("", False),
    ("COLUNAS OBRIGATÓRIAS (azul escuro)", True),
    ("uf_favorecida       → Sigla de 2 letras: AC AL AM AP BA CE DF ES GO MA MG MS MT PA PB PE PI PR RJ RN RO RR RS SC SE SP TO", False),
    ("cnpj_emitente       → Apenas números, 14 dígitos. Zeros à esquerda são obrigatórios (ex: 09595247000172)", False),
    ("razao_social        → Nome completo da empresa", False),
    ("uf_emitente         → UF onde a empresa está registrada (ex: SP)", False),
    ("municipio_emitente  → Código IBGE de 5 dígitos da cidade do emitente (sem os 2 primeiros dígitos do estado)", False),
    ("periodo_referencia  → Competência no formato AAAAMM, ex: 202501 para Janeiro/2025", False),
    ("valor_principal     → Valor do ICMS DIFAL em reais, com ponto decimal. Ex: 1500.00", False),
    ("data_vencimento     → Data no formato AAAA-MM-DD. Ex: 2025-02-20", False),
    ("", False),
    ("COLUNAS OPCIONAIS (azul médio)", True),
    ("detalhamento_receita → Código numérico exigido por algumas UFs. Consulte o portal GNRE para cada UF.", False),
    ("valor_fecp          → FCP/FECP obrigatório em: AL BA CE MA MG PB PE PI RJ SE", False),
    ("documento_origem_tipo → 55=NF-e  57=CT-e  65=NFC-e  10=NF papel. Deixe vazio se a UF não exigir.", False),
    ("documento_origem    → Chave de acesso (44 dígitos para NF-e) ou número do documento", False),
    ("", False),
    ("DICAS", True),
    ("• Formate as células de CNPJ/CEP/Telefone como TEXTO antes de digitar", False),
    ("• data_vencimento e data_pagamento aceitam também o formato DD/MM/AAAA", False),
    ("• Deixe 'receita' em branco — o sistema usa 100080 (DIFAL) automaticamente", False),
    ("• Erros 213/217 do portal: verifique detalhamento_receita e documento_origem para cada UF", False),
]

wi.column_dimensions["A"].width = 100
for i, (texto, negrito) in enumerate(instrucoes, start=1):
    c = wi.cell(row=i, column=1, value=texto)
    if negrito:
        c.font = Font(bold=True, size=11, color="1F4E79")
    else:
        c.font = Font(size=10)
    c.alignment = Alignment(wrap_text=True)
    wi.row_dimensions[i].height = 18 if not negrito else 22

wb.save(saida)
print(f"Modelo gerado: {saida.resolve()}")
print(f"  • {len(CABECALHOS)} colunas ({sum(OBRIGATORIO)} obrigatórias, {sum(not o for o in OBRIGATORIO)} opcionais)")
print("  • Linha 2 contém as descrições de cada campo")
print("  • Aba 'Instruções' com guia completo de preenchimento")
