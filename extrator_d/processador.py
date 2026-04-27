import io
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

BLUE_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

COLS = ["Data", "Descrição", "Valor", "Tipo"]
COL_WIDTHS = [14, 60, 14, 8]


def _extract_cpf_digits(cpf_input: str) -> str | None:
    """
    Extrai os 6 dígitos do meio do CPF e retorna no formato 'DDD.DDD',
    que é como aparece nas descrições de PIX: ***.468.356-**
    """
    digits = re.sub(r"\D", "", cpf_input.strip())
    if len(digits) == 11:
        mid = digits[3:9]           # posições 4–9
    elif len(digits) == 6:
        mid = digits
    elif len(digits) > 6:
        mid = digits[:6]
    else:
        return digits if digits else None
    return f"{mid[:3]}.{mid[3:]}"  # ex: "468.356"


def _matches(descricao: str, nome: str, cpf_digits: str | None) -> bool:
    desc_upper = descricao.upper()
    if nome and nome.strip() and nome.strip().upper() in desc_upper:
        return True
    if cpf_digits and cpf_digits in descricao:
        return True
    return False


def _write_header(ws):
    for col_idx, (col_name, width) in enumerate(zip(COLS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_row(ws, row_num: int, row: pd.Series, highlight: bool):
    values = [row["data"], row["descricao"], row["valor"], row["tipo"]]
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        if highlight:
            cell.fill = BLUE_FILL
        if col_idx == 2:  # descrição
            cell.alignment = Alignment(wrap_text=False)


def processar(file_content: bytes, buscas: list[dict]) -> io.BytesIO:
    """
    Parâmetros:
        file_content: bytes do CSV (data;descricao;valor;tipo, sem cabeçalho)
        buscas: [{"nome": str, "cpf": str}, ...]  — nome e/ou cpf podem ser vazios

    Retorna:
        BytesIO do arquivo Excel com abas:
          - "Todas D": todas as transações tipo D, encontradas destacadas em azul
          - Uma aba por busca com as transações encontradas para aquele CPF/nome
    """
    try:
        text = file_content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = file_content.decode("latin-1")

    df = pd.read_csv(
        io.StringIO(text),
        header=None,
        sep=";",
        names=["data", "descricao", "valor", "tipo"],
        on_bad_lines="skip",
        dtype=str,
    )

    df_d = df[df["tipo"].str.strip() == "D"].copy().reset_index(drop=True)

    # Pré-processa buscas
    buscas_proc = []
    for b in buscas:
        nome = b.get("nome", "").strip()
        cpf_raw = b.get("cpf", "").strip()
        cpf_digits = _extract_cpf_digits(cpf_raw) if cpf_raw else None
        label = nome or cpf_raw or f"Busca {len(buscas_proc)+1}"
        # Nome da aba: máx 31 caracteres, sem caracteres inválidos
        sheet_name = re.sub(r'[\\/*?\[\]:]', '', label)[:31]
        buscas_proc.append({
            "nome": nome,
            "cpf_digits": cpf_digits,
            "label": label,
            "sheet_name": sheet_name,
        })

    # Para cada linha D, calcula quais buscas ela satisfaz
    match_flags = []  # lista de sets com índices das buscas que a linha atende
    for _, row in df_d.iterrows():
        desc = str(row["descricao"])
        hits = set()
        for i, b in enumerate(buscas_proc):
            if _matches(desc, b["nome"], b["cpf_digits"]):
                hits.add(i)
        match_flags.append(hits)

    wb = Workbook()

    # ── Aba "Todas D" ─────────────────────────────────────────────────────────
    ws_all = wb.active
    ws_all.title = "Todas D"
    _write_header(ws_all)
    for i, (_, row) in enumerate(df_d.iterrows()):
        highlight = len(match_flags[i]) > 0
        _write_row(ws_all, i + 2, row, highlight)
    ws_all.freeze_panes = "A2"

    # ── Uma aba por busca ─────────────────────────────────────────────────────
    for b_idx, b in enumerate(buscas_proc):
        ws = wb.create_sheet(title=b["sheet_name"])
        _write_header(ws)
        row_num = 2
        for i, (_, row) in enumerate(df_d.iterrows()):
            if b_idx in match_flags[i]:
                _write_row(ws, row_num, row, highlight=True)
                row_num += 1
        ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
