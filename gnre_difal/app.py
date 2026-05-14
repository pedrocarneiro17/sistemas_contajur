"""
Blueprint GNRE/DIFAL integrado ao sistemas_contajur.
"""

import logging
import os
import tempfile
import threading
import uuid
from pathlib import Path

from flask import Blueprint, jsonify, render_template, request, send_file, make_response
from flask_cors import CORS

from gnre_difal.leitor_dados import ler_planilha
from gnre_difal.xml_builder import (
    construir_envelope_soap_consulta,
    construir_envelope_soap_recepcao,
    construir_lote_xml,
    construir_lote_xml_arquivo,
    dividir_em_lotes,
)
from gnre_difal.exportador import salvar_log_protocolos, salvar_resultados
from gnre_difal.models import RetornoGuia

gnre_difal_bp = Blueprint(
    "gnre_difal",
    __name__,
    url_prefix="/gnre-difal",
    template_folder="templates",
)
CORS(gnre_difal_bp)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Armazena o estado dos jobs em memória
# { job_id: { "status": ..., "log": [...], "resultado_path": ..., "erro": ... } }
JOBS: dict[str, dict] = {}
JOBS_LOCK = threading.Lock()


# ---------------------------------------------------------------------------
# Rotas
# ---------------------------------------------------------------------------

@gnre_difal_bp.route("/")
def index():
    return render_template("index.html")


@gnre_difal_bp.route("/baixar_modelo")
def baixar_modelo():
    """Gera e retorna a planilha modelo em branco com todas as colunas GNRE DIFAL."""
    import io
    import pandas as pd
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # (nome_coluna, obrigatório, descrição/exemplo)
    COLUNAS = [
        ("uf_favorecida",             True,  "UF destinatária — ex: SP"),
        ("cnpj_emitente",             True,  "CNPJ só números — ex: 09595247000172"),
        ("razao_social",              True,  "Razão social da empresa"),
        ("uf_emitente",               True,  "UF da empresa — ex: SP"),
        ("municipio_emitente",        True,  "Código IBGE 5 dígitos — ex: 50308"),
        ("periodo_referencia",        True,  "Competência AAAAMM — ex: 202501"),
        ("valor_principal",           True,  "Valor ICMS DIFAL — ex: 1500.00"),
        ("data_vencimento",           True,  "Vencimento AAAA-MM-DD — ex: 2025-02-20"),
        ("tipo_gnre",                 False, "0=Simples (padrão) | 1=Múlt. docs | 2=Múlt. receitas"),
        ("receita",                   False, "Código 6 dígitos — vazio = 100080 (DIFAL padrão)"),
        ("detalhamento_receita",      False, "Código exigido por certas UFs — consulte o portal"),
        ("valor_fecp",                False, "Valor FCP/FECP — AL BA CE MA MG PB PE PI RJ SE"),
        ("valor_juros",               False, "Juros — se houver"),
        ("valor_multa",               False, "Multa — se houver"),
        ("data_pagamento",            False, "Data pagamento AAAA-MM-DD — vazio = usa vencimento"),
        ("endereco",                  False, "Endereço do emitente"),
        ("cep",                       False, "CEP só números — ex: 01310100"),
        ("telefone",                  False, "Telefone só números — ex: 11999999999"),
        ("ie_emitente",               False, "IE do emitente na UF favorecida (se inscrito)"),
        ("cnpj_destinatario",         False, "CNPJ do destinatário (quando exigido pela UF)"),
        ("cpf_destinatario",          False, "CPF do destinatário"),
        ("ie_destinatario",           False, "IE do destinatário"),
        ("razao_social_destinatario", False, "Razão social do destinatário"),
        ("municipio_destinatario",    False, "Código IBGE 5 dígitos do destinatário"),
        ("documento_origem_tipo",     False, "55=NF-e | 57=CT-e | 65=NFC-e | 10=NF papel"),
        ("documento_origem",          False, "Chave/número do documento de origem"),
        ("identificador_guia",        False, "ID interno até 10 dígitos — para controle próprio"),
    ]

    cabecalhos = [c[0] for c in COLUNAS]

    # DataFrame só com cabeçalho — sem dados de exemplo
    df = pd.DataFrame(columns=cabecalhos)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="DIFAL", index=False)

        ws = writer.sheets["DIFAL"]

        cor_obrig  = "1F4E79"  # azul escuro  → obrigatório
        cor_opcio  = "2E75B6"  # azul médio   → opcional
        cor_desc   = "D9E1F2"  # azul claro   → linha de descrição
        borda = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"),
        )

        # Linha 1 — cabeçalhos coloridos
        for col_idx, (nome, obrig, _) in enumerate(COLUNAS, 1):
            cell = ws.cell(row=1, column=col_idx, value=nome)
            cell.fill      = PatternFill("solid", fgColor=cor_obrig if obrig else cor_opcio)
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = borda

        # Linha 2 — descrições (guia de preenchimento)
        for col_idx, (_, _, descricao) in enumerate(COLUNAS, 1):
            cell = ws.cell(row=2, column=col_idx, value=descricao)
            cell.fill      = PatternFill("solid", fgColor=cor_desc)
            cell.font      = Font(italic=True, color="404040", size=9)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border    = borda

        # Largura das colunas
        for col_idx, (nome, _, descricao) in enumerate(COLUNAS, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                max(len(nome), len(descricao) // 2, 14), 40
            )

        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 36
        ws.freeze_panes = "A3"

        # Aba Instruções
        wi = writer.book.create_sheet("Instruções")
        instrucoes = [
            ("INSTRUÇÕES DE PREENCHIMENTO", True),
            ("", False),
            ("COLUNAS OBRIGATÓRIAS (azul escuro)", True),
            ("uf_favorecida        → Sigla 2 letras: AC AL AM AP BA CE DF ES GO MA MG MS MT PA PB PE PI PR RJ RN RO RR RS SC SE SP TO", False),
            ("cnpj_emitente        → Apenas números, 14 dígitos. Zeros à esquerda obrigatórios. Ex: 09595247000172", False),
            ("razao_social         → Nome completo da empresa", False),
            ("uf_emitente          → UF onde a empresa está registrada. Ex: SP", False),
            ("municipio_emitente   → Código IBGE 5 dígitos sem os 2 primeiros (prefixo do estado). Ex: 50308 = São Paulo", False),
            ("periodo_referencia   → Formato AAAAMM. Ex: 202501 = Janeiro/2025", False),
            ("valor_principal      → Valor do ICMS DIFAL com ponto decimal. Ex: 1500.00", False),
            ("data_vencimento      → Formato AAAA-MM-DD ou DD/MM/AAAA. Ex: 2025-02-20", False),
            ("", False),
            ("COLUNAS OPCIONAIS (azul médio)", True),
            ("detalhamento_receita → Código numérico exigido por algumas UFs. Acesse o portal GNRE, selecione a UF e receita para ver os valores válidos.", False),
            ("valor_fecp           → FCP obrigatório em: AL BA CE MA MG PB PE PI RJ SE", False),
            ("documento_origem_tipo → 55=NF-e  57=CT-e  65=NFC-e  10=NF papel. Deixe vazio se a UF não aceitar (erro 217).", False),
            ("documento_origem     → Chave de acesso (44 dígitos para NF-e) ou número do documento", False),
            ("", False),
            ("DICAS", True),
            ("• Formate colunas de CNPJ/CEP/Telefone como TEXTO antes de digitar", False),
            ("• Erro 213 do portal: preencha detalhamento_receita para a UF indicada", False),
            ("• Erro 217 do portal: deixe documento_origem e documento_origem_tipo em branco para essa UF", False),
            ("• receita pode ficar vazio — o sistema usa 100080 (DIFAL) automaticamente", False),
        ]
        wi.column_dimensions["A"].width = 110
        for i, (texto, negrito) in enumerate(instrucoes, 1):
            c = wi.cell(row=i, column=1, value=texto)
            c.font = Font(bold=True, size=11, color="1F4E79") if negrito else Font(size=10)
            c.alignment = Alignment(wrap_text=True)
            wi.row_dimensions[i].height = 22 if negrito else 18

    buf.seek(0)
    response = make_response(buf.read())
    response.headers["Content-Disposition"] = "attachment; filename=modelo_difal_gnre.xlsx"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response


@gnre_difal_bp.route("/processar", methods=["POST"])
def processar():
    """Recebe os arquivos e configurações, inicia processamento em background."""
    # Valida os uploads
    cert_file = request.files.get("certificado")
    dados_file = request.files.get("planilha")

    if not cert_file or cert_file.filename == "":
        return jsonify({"erro": "Certificado não enviado."}), 400

    ext_cert = Path(cert_file.filename).suffix.lower()
    if ext_cert not in (".pfx", ".p12"):
        return jsonify({
            "erro": f"Arquivo de certificado inválido: '{cert_file.filename}'.\n"
                    f"O campo 'Certificado' aceita apenas arquivos .pfx ou .p12.\n"
                    f"Parece que foi enviado um arquivo {ext_cert or 'sem extensão'} por engano."
        }), 400

    if not dados_file or dados_file.filename == "":
        return jsonify({"erro": "Planilha de dados não enviada."}), 400

    ext_dados = Path(dados_file.filename).suffix.lower()
    if ext_dados not in (".xlsx", ".xls", ".xlsm", ".csv"):
        return jsonify({
            "erro": f"Arquivo de planilha inválido: '{dados_file.filename}'.\n"
                    f"O campo 'Planilha' aceita apenas .xlsx, .xls ou .csv."
        }), 400

    form = request.form
    cnpj = "".join(c for c in form.get("cnpj", "") if c.isdigit())
    ie = form.get("ie", "").strip() or None
    senha_cert = form.get("senha_cert", "")
    dry_run = form.get("dry_run") == "true"

    if len(cnpj) != 14:
        return jsonify({"erro": "CNPJ inválido. Informe os 14 dígitos."}), 400

    # Salva arquivos em temp
    tmp_dir = Path(tempfile.mkdtemp(prefix="gnre_"))
    cert_path = tmp_dir / "certificado.pfx"
    dados_path = tmp_dir / ("dados" + Path(dados_file.filename).suffix)
    saida_dir = tmp_dir / "saida"
    saida_dir.mkdir()

    cert_file.save(str(cert_path))
    dados_file.save(str(dados_path))

    job_id = str(uuid.uuid4())[:8]
    with JOBS_LOCK:
        JOBS[job_id] = {
            "status": "iniciando",
            "log": [],
            "resultado_path": None,
            "protocolo_path": None,
            "erro": None,
            "total_guias": 0,
            "lotes_enviados": 0,
            "total_lotes": 0,
        }

    thread = threading.Thread(
        target=_executar_job,
        args=(job_id, cert_path, str(senha_cert), dados_path, saida_dir, cnpj, ie, dry_run),
        daemon=True,
    )
    thread.start()

    return jsonify({"job_id": job_id})


@gnre_difal_bp.route("/status/<job_id>")
def status(job_id: str):
    with JOBS_LOCK:
        job = JOBS.get(job_id)
    if not job:
        return jsonify({"erro": "Job não encontrado."}), 404
    return jsonify(job)


@gnre_difal_bp.route("/download/<job_id>/<tipo>")
def download(job_id: str, tipo: str):
    with JOBS_LOCK:
        job = JOBS.get(job_id)
    if not job:
        return jsonify({"erro": "Job não encontrado."}), 404

    if tipo == "resultado" and job.get("resultado_path"):
        return send_file(job["resultado_path"], as_attachment=True)
    if tipo == "protocolos" and job.get("protocolo_path"):
        return send_file(job["protocolo_path"], as_attachment=True)
    if tipo == "xmls" and job.get("xmls_zip"):
        return send_file(job["xmls_zip"], as_attachment=True)

    return jsonify({"erro": "Arquivo não disponível."}), 404


# ---------------------------------------------------------------------------
# Lógica de processamento (roda em thread separada)
# ---------------------------------------------------------------------------

def _log(job_id: str, msg: str, nivel: str = "info") -> None:
    logger.info("[%s] %s", job_id, msg)
    with JOBS_LOCK:
        JOBS[job_id]["log"].append({"nivel": nivel, "msg": msg})


def _executar_job(
    job_id: str,
    cert_path: Path,
    senha_cert: str,
    dados_path: Path,
    saida_dir: Path,
    cnpj_emitente: str,
    ie_emitente: str | None,
    dry_run: bool,
) -> None:
    import gnre_difal.config as cfg  # import aqui para não misturar estado global

    try:
        _atualizar(job_id, status="lendo_dados")
        _log(job_id, "Lendo planilha de dados...")

        guias = ler_planilha(dados_path)
        if not guias:
            _atualizar(job_id, status="erro", erro="Nenhuma guia válida encontrada na planilha.")
            return

        # Sobrepõe CNPJ/IE lidos do formulário
        for g in guias:
            if not g.cnpj_emitente:
                g.cnpj_emitente = cnpj_emitente
            if ie_emitente and not g.ie_emitente:
                g.ie_emitente = ie_emitente

        total = len(guias)
        lotes = dividir_em_lotes(guias, tamanho=cfg.MAX_GUIAS_POR_LOTE)
        _atualizar(job_id, total_guias=total, total_lotes=len(lotes))
        _log(job_id, f"{total} guias válidas | {len(lotes)} lote(s)")

        # Salva XMLs localmente
        xmls_dir = saida_dir / "xmls"
        xmls_dir.mkdir()
        for i, lote in enumerate(lotes, 1):
            xml = construir_lote_xml_arquivo(lote)   # inclui <?xml standalone="yes"?>
            (xmls_dir / f"lote_{i:03d}.xml").write_text(xml, encoding="utf-8")

        # Cria zip dos XMLs para download
        import zipfile
        zip_path = saida_dir / "xmls_lotes.zip"
        with zipfile.ZipFile(zip_path, "w") as zf:
            for xml_file in sorted(xmls_dir.glob("*.xml")):
                zf.write(xml_file, xml_file.name)
        _atualizar(job_id, xmls_zip=str(zip_path))

        if dry_run:
            _log(job_id, "Modo dry-run: XMLs gerados, envio cancelado.")
            _atualizar(job_id, status="concluido")
            return

        # --- Envio real ao portal ---
        _atualizar(job_id, status="enviando")

        # Importa aqui para usar o cert informado no form
        from gnre_difal.certificate import limpar_pem_temporarios
        from gnre_difal.soap_client import GnreClient

        # Temporariamente sobrescreve o config de certificado
        cfg.CERT_PATH = str(cert_path)
        cfg.CERT_PASSWORD = senha_cert

        log_protocolos = []
        resultados_totais: list[RetornoGuia] = []

        with GnreClient() as cliente:
            for num_lote, lote in enumerate(lotes, 1):
                _log(job_id, f"Enviando lote {num_lote}/{len(lotes)} ({len(lote)} guias)...")
                try:
                    xml_lote = construir_lote_xml(lote)
                    envelope = construir_envelope_soap_recepcao(xml_lote)
                    retorno = cliente.enviar_lote(envelope)

                    log_protocolos.append({
                        "lote": num_lote,
                        "protocolo": retorno.protocolo,
                        "situacao": retorno.situacao,
                        "mensagem": retorno.mensagem,
                        "guias": len(lote),
                    })

                    if retorno.protocolo:
                        _log(job_id, f"✔ Protocolo recebido: {retorno.protocolo}")
                        # Mantém lista de protocolos visível no job para a UI
                        with JOBS_LOCK:
                            if "protocolos_recebidos" not in JOBS[job_id]:
                                JOBS[job_id]["protocolos_recebidos"] = []
                            JOBS[job_id]["protocolos_recebidos"].append(retorno.protocolo)

                        envelope_consulta = construir_envelope_soap_consulta(retorno.protocolo)
                        guias_ret = cliente.consultar_com_retry(envelope_consulta)
                        resultados_totais.extend(guias_ret)
                        _log(job_id, f"Lote {num_lote}: {len(guias_ret)} guias processadas.")
                    else:
                        _log(job_id,
                             f"Lote {num_lote}: protocolo não retornado. "
                             f"Situação='{retorno.situacao}' Msg='{retorno.mensagem}'. "
                             f"Verifique saida/debug/ para ver a resposta bruta do portal.",
                             "aviso")

                    _atualizar(job_id, lotes_enviados=num_lote)

                except Exception as e:
                    _log(job_id, f"Erro no lote {num_lote}: {e}", "erro")
                    log_protocolos.append({
                        "lote": num_lote, "protocolo": "",
                        "situacao": "ERRO", "mensagem": str(e), "guias": len(lote),
                    })

        # Exporta resultados
        _log(job_id, "Salvando resultados...")
        resultado_path = salvar_resultados(guias, resultados_totais, saida_dir)
        protocolo_path = salvar_log_protocolos(log_protocolos, saida_dir)

        _atualizar(
            job_id,
            status="concluido",
            resultado_path=str(resultado_path),
            protocolo_path=str(protocolo_path),
        )
        _log(job_id, "Processamento concluído com sucesso!")

    except Exception as e:
        logger.exception("[%s] Erro fatal", job_id)
        _atualizar(job_id, status="erro", erro=str(e))
        _log(job_id, f"Erro fatal: {e}", "erro")


def _atualizar(job_id: str, **kwargs) -> None:
    with JOBS_LOCK:
        if job_id in JOBS:
            JOBS[job_id].update(kwargs)


# ---------------------------------------------------------------------------
